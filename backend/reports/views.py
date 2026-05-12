"""Exportes de inventario y ventas POS (JSON, Excel, PDF)."""

from __future__ import annotations

from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum, Value
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from PIL import Image as PILImage
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .pdf_brand import BOUTIQUE_PDF_PAGE_SIZE, pdf_header_image_bytes
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from inventory.models import InventoryItem
from inventory.unit_hierarchy import hierarchy_label, split_stock_hierarchy
from pos.models import Sale, SaleLine
from branches.models import Branch
from suppliers.models import PurchaseOrder, Supplier

_REPORT_KINDS = frozenset({'json', 'pdf', 'xlsx'})


def _set_no_store(resp):
    resp['Cache-Control'] = 'no-store, private'
    return resp


def _add_logo_to_ws(ws, cell: str = 'A1', width_px: int = 180) -> None:
    """Inserta logo de marca en una hoja de Excel, si existe."""
    logo_buf = pdf_header_image_bytes()
    try:
        with PILImage.open(logo_buf) as pil_im:
            src_w, src_h = pil_im.size
            if src_w <= 0 or src_h <= 0:
                return
            new_h = max(24, int((src_h * width_px) / src_w))
            resized = pil_im.convert('RGB').resize((width_px, new_h))
            out = BytesIO()
            resized.save(out, format='PNG')
            out.seek(0)
    except Exception:
        return
    img = XLImage(out)
    ws.add_image(img, cell)


def _report_export_format(request) -> str:
    """
    Tipo de salida del reporte (json, xlsx, pdf).

    - Cabecera `X-Boutique-Report` tiene prioridad (no depende del query; algunos
      proxies o caches alteran query strings).
    - No usar query `format`: REST framework lo reserva.
    - Query: `out`, `tipo`, `export`.
    """
    hdr = (request.headers.get('X-Boutique-Report') or '').strip().lower()
    if hdr in _REPORT_KINDS:
        return hdr
    raw = (
        request.query_params.get('tipo')
        or request.query_params.get('export')
        or request.query_params.get('out')
        or 'json'
    )
    return raw.lower()


def _parse_branch(request) -> int | None:
    raw = request.query_params.get('branch')
    if raw is None or raw == '':
        return None
    try:
        n = int(raw)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _parse_inventory_scope(request) -> str | None:
    raw = (request.query_params.get('scope') or request.query_params.get('ubicacion') or '').strip().lower()
    return raw if raw in {'tienda', 'b1', 'b2', 'b3'} else None


def _bodega_branch_ids() -> dict[str, int | None]:
    by_name = {b.name.strip().lower(): int(b.id) for b in Branch.objects.filter(is_active=True)}
    return {
        'b1': by_name.get('bodega 1'),
        'b2': by_name.get('bodega 2'),
        'b3': by_name.get('bodega 3'),
    }


def _inventory_rows(branch_id: int | None, scope: str | None = None) -> list[dict]:
    qs = InventoryItem.objects.select_related('branch', 'category').order_by('branch__name', 'line', 'sku')
    if branch_id is not None:
        qs = qs.filter(branch_id=branch_id)
    elif scope is not None:
        bodega_ids_map = _bodega_branch_ids()
        bodega_ids = [bid for bid in bodega_ids_map.values() if isinstance(bid, int) and bid > 0]
        if scope == 'tienda':
            if bodega_ids:
                qs = qs.exclude(branch_id__in=bodega_ids)
        else:
            bid = bodega_ids_map.get(scope)
            if isinstance(bid, int) and bid > 0:
                qs = qs.filter(branch_id=bid)
            else:
                qs = qs.none()
    rows = []
    for obj in qs:
        upp = int(obj.units_per_package)
        ppf = int(obj.packages_per_fardo)
        units_per_fardo = upp * ppf
        rows.append(
            {
                'id': obj.pk,
                'nombre': obj.name,
                'branch_id': obj.branch_id,
                'branch_name': obj.branch.name if obj.branch_id else '',
                'categoria': obj.category.name if obj.category_id else '',
                'units_per_package': obj.units_per_package,
                'units_per_fardo': units_per_fardo,
                'cantidad': obj.quantity,
                'precio_unitario': str(obj.unit_price),
                'precio_costo': str(obj.cost_price),
            }
        )
    return rows


def _pos_sales_rows(branch_id: int | None, limit: int = 500) -> list[dict]:
    qs = Sale.objects.select_related('branch').prefetch_related('lines__inventory_item').order_by('-created_at')
    if branch_id is not None:
        qs = qs.filter(branch_id=branch_id)
    out = []
    for sale in qs[:limit]:
        lines = []
        for ln in sale.lines.all():
            item = ln.inventory_item
            f, p, u = split_stock_hierarchy(ln.quantity, item.units_per_package, item.packages_per_fardo)
            lines.append(
                {
                    'sku': item.sku,
                    'producto': item.name,
                    'cantidad': ln.quantity,
                    'units_per_package': item.units_per_package,
                    'packages_per_fardo': item.packages_per_fardo,
                    'venta_fardos': f,
                    'venta_paquetes': p,
                    'venta_unidades_resto': u,
                    'jerarquia_txt': hierarchy_label(f, p, u),
                    'precio_unitario': str(ln.unit_price),
                    'subtotal': str(Decimal(ln.unit_price) * ln.quantity),
                }
            )
        out.append(
            {
                'id': sale.pk,
                'ubicacion': sale.branch.name,
                'metodo_pago': sale.get_payment_method_display(),
                'total': str(sale.total),
                'fecha': sale.created_at.isoformat(),
                'lineas': lines,
            }
        )
    return out


def _branded_table_pdf(
    document_title: str,
    report_heading: str,
    headers: list[str],
    data_rows: list[list],
    filename: str,
) -> HttpResponse:
    """PDF en carta vertical con logo de empresa (o cabecera de respaldo) y tabla."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=BOUTIQUE_PDF_PAGE_SIZE,
        title=document_title,
        leftMargin=28,
        rightMargin=28,
        topMargin=34,
        bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    title_left = ParagraphStyle('repTitleLeft', parent=styles['Title'], alignment=TA_LEFT)
    normal_left = ParagraphStyle('repNormalLeft', parent=styles['Normal'], alignment=TA_LEFT)
    story: list = []

    logo_buf = pdf_header_image_bytes()
    logo_buf.seek(0)
    with PILImage.open(logo_buf) as pil_im:
        lw, lh = pil_im.size
    logo_buf.seek(0)
    # Logo compacto, esquina superior izquierda (ancho máx. ~41 mm en puntos).
    logo_max_w = 112.0
    display_w = min(logo_max_w, float(lw)) if lw else logo_max_w
    display_h = display_w * (float(lh) / float(lw)) if lw else 32.0
    story.append(PdfImage(logo_buf, width=display_w, height=display_h, hAlign='LEFT'))
    story.append(Spacer(1, 8))

    story.append(Paragraph(f'<b>{report_heading}</b>', title_left))
    story.append(
        Paragraph(
            f'Generado: {timezone.now().strftime("%Y-%m-%d %H:%M")} — Aluminios Ixmucane',
            normal_left,
        )
    )
    story.append(Spacer(1, 12))

    table_data = [headers] + data_rows
    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c40000')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return _set_no_store(resp)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def inventory_report(request, salida: str | None = None):
    if salida is not None and str(salida).strip() != '':
        fmt = str(salida).strip().lower()
    else:
        fmt = _report_export_format(request)
    if fmt not in _REPORT_KINDS:
        return _set_no_store(
            JsonResponse(
                {
                    'detail': f'Formato no valido ({fmt}). Use /reports/inventario/json/, /pdf/, /xlsx/ o query out= (o tipo=).',
                },
                status=400,
            )
        )
    branch_id = _parse_branch(request)
    scope = _parse_inventory_scope(request)

    if fmt == 'json':
        return _set_no_store(
            JsonResponse({'generated_at': timezone.now().isoformat(), 'items': _inventory_rows(branch_id, scope)})
        )

    rows = _inventory_rows(branch_id, scope)
    headers = [
        'Nombre',
        'Categoría',
        'U/paquete',
        'U/fardo',
        'Unidades',
        'Precio costo',
        'Precio venta',
    ]

    if fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventario'
        _add_logo_to_ws(ws, 'A1')
        header_row = 8
        bold = Font(bold=True)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = bold
        for r, item in enumerate(rows, start=header_row + 1):
            ws.cell(row=r, column=1, value=item['nombre'])
            ws.cell(row=r, column=2, value=item['categoria'])
            ws.cell(row=r, column=3, value=item['units_per_package'])
            ws.cell(row=r, column=4, value=item['units_per_fardo'])
            ws.cell(row=r, column=5, value=item['cantidad'])
            ws.cell(row=r, column=6, value=item['precio_costo'])
            ws.cell(row=r, column=7, value=item['precio_unitario'])
        buf = BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        buf.close()
        resp = HttpResponse(
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="reporte_inventario_{datetime.now():%Y%m%d_%H%M}.xlsx"'
        return _set_no_store(resp)

    if fmt == 'pdf':
        data_rows = [
            [
                item['nombre'][:48],
                item['categoria'][:24],
                str(item['units_per_package']),
                str(item['units_per_fardo']),
                str(item['cantidad']),
                item['precio_costo'],
                item['precio_unitario'],
            ]
            for item in rows[:400]
        ]
        return _branded_table_pdf(
            'Inventario — Aluminios Ixmucane',
            'Reporte de inventario general',
            headers,
            data_rows,
            f'reporte_inventario_{datetime.now():%Y%m%d_%H%M}.pdf',
        )

    return _set_no_store(
        JsonResponse({'detail': 'Use /reports/inventario/json/, /pdf/, /xlsx/ o query out= (o tipo=).'}, status=400)
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pos_sales_report(request, salida: str | None = None):
    if salida is not None and str(salida).strip() != '':
        fmt = str(salida).strip().lower()
    else:
        fmt = _report_export_format(request)
    if fmt not in _REPORT_KINDS:
        return _set_no_store(
            JsonResponse(
                {
                    'detail': f'Formato no valido ({fmt}). Use /reports/sistema-pos/json/, /pdf/, /xlsx/ o query out= (o tipo=).',
                },
                status=400,
            )
        )
    branch_id = _parse_branch(request)

    if fmt == 'json':
        return _set_no_store(
            JsonResponse({'generated_at': timezone.now().isoformat(), 'ventas': _pos_sales_rows(branch_id)})
        )

    rows = _pos_sales_rows(branch_id)

    if fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ventas'
        _add_logo_to_ws(ws, 'A1')
        h = ['Ticket', 'Fecha', 'Ubicación', 'Pago', 'Total']
        header_row = 8
        bold = Font(bold=True)
        for col, title in enumerate(h, start=1):
            ws.cell(row=header_row, column=col, value=title).font = bold
        r = header_row + 1
        for v in rows:
            ws.cell(row=r, column=1, value=v['id'])
            ws.cell(row=r, column=2, value=v['fecha'][:19])
            ws.cell(row=r, column=3, value=v['ubicacion'])
            ws.cell(row=r, column=4, value=v['metodo_pago'])
            ws.cell(row=r, column=5, value=v['total'])
            r += 1
        ws2 = wb.create_sheet('Detalle líneas')
        _add_logo_to_ws(ws2, 'A1')
        h2 = ['Ticket', 'SKU', 'Producto', 'Cantidad (u)', 'Desglose', 'P.U.', 'Subtotal']
        header_row_2 = 8
        for col, title in enumerate(h2, start=1):
            ws2.cell(row=header_row_2, column=col, value=title).font = bold
        r2 = header_row_2 + 1
        for v in rows:
            for ln in v['lineas']:
                ws2.cell(row=r2, column=1, value=v['id'])
                ws2.cell(row=r2, column=2, value=ln['sku'])
                ws2.cell(row=r2, column=3, value=ln['producto'][:60])
                ws2.cell(row=r2, column=4, value=ln['cantidad'])
                ws2.cell(row=r2, column=5, value=ln['jerarquia_txt'])
                ws2.cell(row=r2, column=6, value=ln['precio_unitario'])
                ws2.cell(row=r2, column=7, value=ln['subtotal'])
                r2 += 1
        buf = BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        buf.close()
        resp = HttpResponse(
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="reporte_pos_{datetime.now():%Y%m%d_%H%M}.xlsx"'
        return _set_no_store(resp)

    if fmt == 'pdf':
        data_rows = []
        for v in rows[:200]:
            data_rows.append([str(v['id']), v['fecha'][:19], v['metodo_pago'], v['total']])
        return _branded_table_pdf(
            'Ventas POS — Aluminios Ixmucane',
            'Reporte de ventas POS (tickets)',
            ['Ticket', 'Fecha', 'Pago', 'Total'],
            data_rows,
            f'reporte_pos_{datetime.now():%Y%m%d_%H%M}.pdf',
        )

    return _set_no_store(
        JsonResponse({'detail': 'Use /reports/sistema-pos/json/, /pdf/, /xlsx/ o query out= (o tipo=).'}, status=400)
    )


def _suppliers_rows() -> list[dict]:
    suppliers = Supplier.objects.prefetch_related('purchase_orders__lines').order_by('name', 'id')
    out = []
    for s in suppliers:
        orders = list(s.purchase_orders.all())
        total_lines = sum(o.lines.count() for o in orders)
        out.append({
            'id': s.pk,
            'nombre': str(s) ,
            'razon_social': s.razon_social or '',
            'nit': s.nit or '',
            'contacto': s.contact or '',
            'total_ordenes': len(orders),
            'total_lineas': total_lines,
        })
    return out


def _purchase_orders_rows() -> list[dict]:
    orders = (
        PurchaseOrder.objects
        .select_related('supplier', 'branch')
        .prefetch_related('lines__inventory_item')
        .order_by('-created_at')
    )
    out = []
    for o in orders:
        out.append({
            'id': o.pk,
            'proveedor': str(o.supplier),
            'referencia': o.reference or '',
            'fecha': o.created_at.isoformat(),
            'lineas': o.lines.count(),
        })
    return out


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def suppliers_report(request, salida: str | None = None):
    if salida is not None and str(salida).strip() != '':
        fmt = str(salida).strip().lower()
    else:
        fmt = _report_export_format(request)
    if fmt not in _REPORT_KINDS:
        return _set_no_store(
            JsonResponse({'detail': f'Formato no valido ({fmt}).'}, status=400)
        )

    if fmt == 'json':
        return _set_no_store(
            JsonResponse({
                'generated_at': timezone.now().isoformat(),
                'proveedores': _suppliers_rows(),
                'ordenes': _purchase_orders_rows(),
            })
        )

    s_rows = _suppliers_rows()
    o_rows = _purchase_orders_rows()

    if fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Proveedores'
        _add_logo_to_ws(ws, 'A1')
        h = ['ID', 'Nombre / Razón social', 'NIT', 'Contacto', 'Órdenes', 'Líneas totales']
        bold = Font(bold=True)
        header_row = 8
        for col, title in enumerate(h, start=1):
            ws.cell(row=header_row, column=col, value=title).font = bold
        for r, row in enumerate(s_rows, start=header_row + 1):
            ws.cell(row=r, column=1, value=row['id'])
            ws.cell(row=r, column=2, value=row['nombre'])
            ws.cell(row=r, column=3, value=row['nit'])
            ws.cell(row=r, column=4, value=row['contacto'])
            ws.cell(row=r, column=5, value=row['total_ordenes'])
            ws.cell(row=r, column=6, value=row['total_lineas'])

        ws2 = wb.create_sheet('Órdenes de compra')
        _add_logo_to_ws(ws2, 'A1')
        h2 = ['ID', 'Proveedor', 'Referencia', 'Fecha', 'Líneas']
        for col, title in enumerate(h2, start=1):
            ws2.cell(row=header_row, column=col, value=title).font = bold
        for r, row in enumerate(o_rows, start=header_row + 1):
            ws2.cell(row=r, column=1, value=row['id'])
            ws2.cell(row=r, column=2, value=row['proveedor'])
            ws2.cell(row=r, column=3, value=row['referencia'])
            ws2.cell(row=r, column=4, value=row['fecha'][:19])
            ws2.cell(row=r, column=5, value=row['lineas'])

        buf = BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        buf.close()
        resp = HttpResponse(
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="reporte_proveedores_{datetime.now():%Y%m%d_%H%M}.xlsx"'
        return _set_no_store(resp)

    if fmt == 'pdf':
        data_rows = [
            [str(r['id']), r['nombre'][:40], r['nit'][:20], r['contacto'][:30], str(r['total_ordenes'])]
            for r in s_rows[:300]
        ]
        return _branded_table_pdf(
            'Proveedores — Aluminios Ixmucane',
            'Reporte de proveedores',
            ['ID', 'Nombre / Razón social', 'NIT', 'Contacto', 'Órdenes'],
            data_rows,
            f'reporte_proveedores_{datetime.now():%Y%m%d_%H%M}.pdf',
        )

    return _set_no_store(
        JsonResponse({'detail': 'Use json, pdf o xlsx.'}, status=400)
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cobros_report(request, salida: str = 'pdf'):
    """Reporte de cuentas por cobrar (ventas a crédito o pago pendiente)."""
    fmt = salida if salida in _REPORT_KINDS else _report_export_format(request)

    qs = (
        Sale.objects.select_related('branch')
        .prefetch_related('lines__inventory_item')
        .filter(payment_status__in=['credit', 'pending'])
        .order_by('created_at')
    )
    rows = []
    for sale in qs:
        rows.append({
            'id': sale.pk,
            'fecha': sale.created_at.strftime('%Y-%m-%d %H:%M'),
            'cliente': sale.customer_name or 'Consumidor final',
            'telefono': sale.customer_phone or '',
            'metodo': sale.get_payment_method_display(),
            'estado': sale.get_payment_status_display(),
            'plazo_dias': sale.credit_days,
            'nota': sale.credit_note,
            'total': str(sale.total),
        })

    if fmt == 'json':
        return _set_no_store(JsonResponse({'results': rows, 'count': len(rows)}))

    if fmt == 'pdf':
        from decimal import Decimal as D
        total_pendiente = sum(D(r['total']) for r in rows)
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=BOUTIQUE_PDF_PAGE_SIZE, title='Cuentas por Cobrar')
        styles = getSampleStyleSheet()
        heading_left = ParagraphStyle('cobHeading', parent=styles['Heading2'], alignment=TA_LEFT)
        normal_left  = ParagraphStyle('cobNormal',  parent=styles['Normal'],   alignment=TA_LEFT)
        h3_left      = ParagraphStyle('cobH3',      parent=styles['Heading3'], alignment=TA_LEFT)
        italic_left  = ParagraphStyle('cobItalic',  parent=styles['Italic'],   alignment=TA_LEFT)

        logo_buf = pdf_header_image_bytes()
        logo_buf.seek(0)
        with PILImage.open(logo_buf) as pil_im:
            lw, lh = pil_im.size
        logo_buf.seek(0)
        logo_max_w = 96.0
        logo_w = min(logo_max_w, float(lw)) if lw else logo_max_w
        logo_h = logo_w * (float(lh) / float(lw)) if lw else 28.0

        story = [
            PdfImage(logo_buf, width=logo_w, height=logo_h, hAlign='LEFT'),
            Spacer(1, 8),
            Paragraph('<b>Reporte de Cuentas por Cobrar</b>', heading_left),
            Paragraph(f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}', normal_left),
            Spacer(1, 12),
        ]
        data = [['Ticket', 'Fecha', 'Cliente', 'Teléfono', 'Estado', 'Plazo', 'Total']]
        for r in rows:
            data.append([
                f'#{r["id"]}',
                r['fecha'],
                r['cliente'][:30],
                r['telefono'][:16],
                r['estado'],
                f'{r["plazo_dias"]}d' if r['plazo_dias'] else '—',
                f'Q {r["total"]}',
            ])
        tbl = Table(data, repeatRows=1, colWidths=[44, 88, 130, 80, 64, 36, 72])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFF')]),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 16))
        story.append(Paragraph(f'<b>Total pendiente de cobro: Q {total_pendiente:.2f}</b>', h3_left))
        story.append(Spacer(1, 24))
        story.append(Paragraph(
            'Este reporte incluye únicamente ventas con estado "Crédito" o "Pago pendiente".',
            italic_left,
        ))
        doc.build(story)
        pdf = buf.getvalue()
        buf.close()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="cuentas_por_cobrar_{datetime.now():%Y%m%d_%H%M}.pdf"'
        return _set_no_store(resp)

    return _set_no_store(JsonResponse({'detail': 'Use json o pdf.'}, status=400))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  REPORTE DE GANANCIAS (semana / quincena / mes)                          ║
# ╚══════════════════════════════════════════════════════════════════════════╝

_PERIODOS_GANANCIAS = {
    'semana': (7, 'Última semana (7 días)'),
    'quincena': (15, 'Última quincena (15 días)'),
    'mes': (30, 'Último mes (30 días)'),
}


def _zero_dec(places: int = 2) -> Value:
    return Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=places))


def _line_margin_expr() -> ExpressionWrapper:
    """`quantity * (unit_price - cost_price)` — ganancia bruta de cada línea."""
    cost_f = Coalesce(
        F('inventory_item__cost_price'),
        _zero_dec(2),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    return ExpressionWrapper(
        F('quantity') * (F('unit_price') - cost_f),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )


def _line_revenue_expr() -> ExpressionWrapper:
    return ExpressionWrapper(
        F('quantity') * F('unit_price'),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )


def _line_cost_expr() -> ExpressionWrapper:
    cost_f = Coalesce(
        F('inventory_item__cost_price'),
        _zero_dec(2),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    return ExpressionWrapper(
        F('quantity') * cost_f,
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )


def _q(d: Decimal | None) -> str:
    return str((d or Decimal('0')).quantize(Decimal('0.01')))


def _margen_pct(ganancia: Decimal, ventas: Decimal) -> str:
    if not ventas or ventas == 0:
        return '0.00'
    return str((ganancia / ventas * Decimal('100')).quantize(Decimal('0.01')))


def _ganancias_data(periodo: str, branch_id: int | None) -> dict:
    """Calcula KPIs y desgloses para el periodo dado."""
    days, label = _PERIODOS_GANANCIAS[periodo]
    now = timezone.now()
    since = now - timedelta(days=days)
    prev_since = since - timedelta(days=days)

    sales_qs = Sale.objects.filter(created_at__gte=since)
    if branch_id is not None:
        sales_qs = sales_qs.filter(branch_id=branch_id)

    lines_qs = SaleLine.objects.filter(sale__created_at__gte=since)
    if branch_id is not None:
        lines_qs = lines_qs.filter(sale__branch_id=branch_id)

    revenue_expr = _line_revenue_expr()
    cost_expr = _line_cost_expr()
    margin_expr = _line_margin_expr()

    # ── KPIs globales del periodo
    line_totals = lines_qs.aggregate(
        ingresos=Sum(revenue_expr),
        costo=Sum(cost_expr),
        ganancia=Sum(margin_expr),
        unidades=Sum('quantity'),
    )
    ingresos = line_totals['ingresos'] or Decimal('0')
    costo = line_totals['costo'] or Decimal('0')
    ganancia = line_totals['ganancia'] or Decimal('0')
    unidades = int(line_totals['unidades'] or 0)

    sale_totals = sales_qs.aggregate(
        tickets=Count('id'),
        descuentos=Sum('discount'),
        ventas_brutas=Sum('total'),
    )
    tickets = int(sale_totals['tickets'] or 0)
    descuentos = sale_totals['descuentos'] or Decimal('0')
    ventas_brutas = sale_totals['ventas_brutas'] or Decimal('0')

    # ganancia neta = ganancia bruta - descuentos (los descuentos salen del margen)
    ganancia_neta = ganancia - descuentos
    ticket_promedio = (ventas_brutas / tickets) if tickets > 0 else Decimal('0')
    ganancia_promedio = (ganancia_neta / tickets) if tickets > 0 else Decimal('0')

    # ── Comparación con periodo anterior (mismo número de días previos)
    prev_lines_qs = SaleLine.objects.filter(
        sale__created_at__gte=prev_since,
        sale__created_at__lt=since,
    )
    prev_sales_qs = Sale.objects.filter(
        created_at__gte=prev_since, created_at__lt=since,
    )
    if branch_id is not None:
        prev_lines_qs = prev_lines_qs.filter(sale__branch_id=branch_id)
        prev_sales_qs = prev_sales_qs.filter(branch_id=branch_id)
    prev_totals = prev_lines_qs.aggregate(
        ingresos=Sum(revenue_expr),
        ganancia=Sum(margin_expr),
    )
    prev_sales_totals = prev_sales_qs.aggregate(
        tickets=Count('id'),
        descuentos=Sum('discount'),
        ventas_brutas=Sum('total'),
    )
    prev_ingresos = prev_totals['ingresos'] or Decimal('0')
    prev_ganancia = (prev_totals['ganancia'] or Decimal('0')) - (
        prev_sales_totals['descuentos'] or Decimal('0')
    )
    prev_ventas_brutas = prev_sales_totals['ventas_brutas'] or Decimal('0')
    prev_tickets = int(prev_sales_totals['tickets'] or 0)

    def _delta_pct(curr: Decimal, prev: Decimal) -> str:
        if not prev or prev == 0:
            return '0.00' if not curr or curr == 0 else '100.00'
        return str(((curr - prev) / prev * Decimal('100')).quantize(Decimal('0.01')))

    # ── Serie diaria
    daily_lines = (
        lines_qs.annotate(day=TruncDate('sale__created_at'))
        .values('day')
        .annotate(
            ingresos=Sum(revenue_expr),
            costo=Sum(cost_expr),
            ganancia=Sum(margin_expr),
            unidades=Sum('quantity'),
        )
        .order_by('day')
    )
    daily_sales = (
        sales_qs.annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(
            tickets=Count('id'),
            descuentos=Sum('discount'),
            ventas_brutas=Sum('total'),
        )
        .order_by('day')
    )
    daily_sales_map = {r['day']: r for r in daily_sales}
    daily = []
    for r in daily_lines:
        ds = daily_sales_map.get(r['day']) or {}
        d_descuentos = ds.get('descuentos') or Decimal('0')
        d_ganancia = (r.get('ganancia') or Decimal('0')) - d_descuentos
        d_ingresos = r.get('ingresos') or Decimal('0')
        daily.append({
            'fecha': r['day'].isoformat() if r['day'] else None,
            'ingresos': _q(d_ingresos),
            'ventas': _q(ds.get('ventas_brutas') or d_ingresos),
            'costo': _q(r.get('costo')),
            'descuento': _q(d_descuentos),
            'ganancia': _q(d_ganancia),
            'tickets': int(ds.get('tickets') or 0),
            'unidades': int(r.get('unidades') or 0),
            'margen_pct': _margen_pct(d_ganancia, d_ingresos),
        })

    # ── Top productos por ganancia
    top_productos_rows = (
        lines_qs.values(
            'inventory_item_id',
            'inventory_item__name',
            'inventory_item__sku',
            'inventory_item__category__name',
        )
        .annotate(
            unidades=Sum('quantity'),
            ingresos=Sum(revenue_expr),
            costo=Sum(cost_expr),
            ganancia=Sum(margin_expr),
            apariciones=Count('id'),
        )
        .order_by('-ganancia')[:25]
    )
    top_productos = [
        {
            'id': r['inventory_item_id'],
            'sku': r['inventory_item__sku'] or '',
            'nombre': r['inventory_item__name'] or '',
            'categoria': r['inventory_item__category__name'] or '',
            'unidades': int(r['unidades'] or 0),
            'ingresos': _q(r['ingresos']),
            'costo': _q(r['costo']),
            'ganancia': _q(r['ganancia']),
            'margen_pct': _margen_pct(r['ganancia'] or Decimal('0'), r['ingresos'] or Decimal('0')),
            'tickets': int(r['apariciones'] or 0),
        }
        for r in top_productos_rows
    ]

    # ── Top categorías
    top_categorias_rows = (
        lines_qs.values('inventory_item__category__name')
        .annotate(
            ingresos=Sum(revenue_expr),
            costo=Sum(cost_expr),
            ganancia=Sum(margin_expr),
            unidades=Sum('quantity'),
            productos=Count('inventory_item_id', distinct=True),
        )
        .order_by('-ganancia')
    )
    top_categorias = [
        {
            'categoria': r['inventory_item__category__name'] or 'Sin categoría',
            'ingresos': _q(r['ingresos']),
            'costo': _q(r['costo']),
            'ganancia': _q(r['ganancia']),
            'margen_pct': _margen_pct(r['ganancia'] or Decimal('0'), r['ingresos'] or Decimal('0')),
            'unidades': int(r['unidades'] or 0),
            'productos': int(r['productos'] or 0),
        }
        for r in top_categorias_rows[:15]
    ]

    # ── Por sucursal
    por_sucursal_rows = (
        lines_qs.values('sale__branch_id', 'sale__branch__name')
        .annotate(
            ingresos=Sum(revenue_expr),
            costo=Sum(cost_expr),
            ganancia=Sum(margin_expr),
            unidades=Sum('quantity'),
            tickets=Count('sale_id', distinct=True),
        )
        .order_by('-ganancia')
    )
    por_sucursal = [
        {
            'branch_id': r['sale__branch_id'],
            'branch_name': r['sale__branch__name'] or '',
            'ingresos': _q(r['ingresos']),
            'costo': _q(r['costo']),
            'ganancia': _q(r['ganancia']),
            'margen_pct': _margen_pct(r['ganancia'] or Decimal('0'), r['ingresos'] or Decimal('0')),
            'unidades': int(r['unidades'] or 0),
            'tickets': int(r['tickets'] or 0),
        }
        for r in por_sucursal_rows
    ]

    # ── Por método de pago
    pago_rows = (
        sales_qs.values('payment_method')
        .annotate(
            tickets=Count('id'),
            ventas=Sum('total'),
        )
        .order_by('-ventas')
    )
    payment_labels = {
        'cash': 'Efectivo',
        'card': 'Tarjeta',
        'other': 'Otro',
    }
    por_pago = [
        {
            'metodo': r['payment_method'],
            'metodo_label': payment_labels.get(r['payment_method'], r['payment_method']),
            'tickets': int(r['tickets'] or 0),
            'ventas': _q(r['ventas']),
        }
        for r in pago_rows
    ]

    # ── Por estado de pago (pagado vs crédito vs pendiente)
    estado_rows = (
        sales_qs.values('payment_status')
        .annotate(
            tickets=Count('id'),
            ventas=Sum('total'),
            cobrado=Sum('amount_paid'),
        )
        .order_by('-ventas')
    )
    estado_labels = {
        'paid': 'Pagado',
        'credit': 'Crédito',
        'pending': 'Pendiente',
    }
    por_estado = [
        {
            'estado': r['payment_status'],
            'estado_label': estado_labels.get(r['payment_status'], r['payment_status']),
            'tickets': int(r['tickets'] or 0),
            'ventas': _q(r['ventas']),
            'cobrado': _q(r['cobrado']),
            'pendiente': _q((r['ventas'] or Decimal('0')) - (r['cobrado'] or Decimal('0'))),
        }
        for r in estado_rows
    ]

    # ── Top clientes (por ganancia)
    top_clientes_rows = (
        lines_qs.values('sale__customer_id', 'sale__customer_name', 'sale__customer_nit')
        .annotate(
            ingresos=Sum(revenue_expr),
            ganancia=Sum(margin_expr),
            tickets=Count('sale_id', distinct=True),
            unidades=Sum('quantity'),
        )
        .order_by('-ganancia')[:15]
    )
    top_clientes = [
        {
            'customer_id': r['sale__customer_id'],
            'nombre': r['sale__customer_name'] or 'Consumidor final',
            'nit': r['sale__customer_nit'] or '',
            'tickets': int(r['tickets'] or 0),
            'unidades': int(r['unidades'] or 0),
            'ingresos': _q(r['ingresos']),
            'ganancia': _q(r['ganancia']),
        }
        for r in top_clientes_rows
    ]

    # ── Tickets más rentables
    margin_per_sale = ExpressionWrapper(
        F('lines__quantity') * (
            F('lines__unit_price')
            - Coalesce(
                F('lines__inventory_item__cost_price'),
                _zero_dec(2),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        ),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )
    top_tickets_rows = (
        sales_qs.annotate(
            ganancia_bruta=Sum(margin_per_sale),
            ingresos_linea=Sum(
                ExpressionWrapper(
                    F('lines__quantity') * F('lines__unit_price'),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            ),
            lines_count=Count('lines', distinct=True),
        )
        .order_by('-ganancia_bruta')[:15]
    )
    top_tickets = []
    for s in top_tickets_rows:
        gb = s.ganancia_bruta or Decimal('0')
        gn = gb - (s.discount or Decimal('0'))
        top_tickets.append({
            'id': s.id,
            'fecha': s.created_at.isoformat(),
            'cliente': s.customer_name or 'Consumidor final',
            'sucursal': s.branch.name if s.branch_id else '',
            'tickets': 1,
            'lineas': int(s.lines_count or 0),
            'ingresos': _q(s.ingresos_linea),
            'descuento': _q(s.discount),
            'total': _q(s.total),
            'ganancia': _q(gn),
        })

    return {
        'periodo': periodo,
        'periodo_label': label,
        'desde': since.isoformat(),
        'hasta': now.isoformat(),
        'branch_id': branch_id,
        'kpis': {
            'tickets': tickets,
            'unidades': unidades,
            'ventas_brutas': _q(ventas_brutas),
            'ingresos': _q(ingresos),
            'descuentos': _q(descuentos),
            'costo': _q(costo),
            'ganancia_bruta': _q(ganancia),
            'ganancia_neta': _q(ganancia_neta),
            'margen_pct': _margen_pct(ganancia_neta, ingresos),
            'ticket_promedio': _q(ticket_promedio),
            'ganancia_promedio': _q(ganancia_promedio),
        },
        'comparacion': {
            'tickets_prev': prev_tickets,
            'ventas_prev': _q(prev_ventas_brutas),
            'ganancia_prev': _q(prev_ganancia),
            'delta_ventas_pct': _delta_pct(ventas_brutas, prev_ventas_brutas),
            'delta_ganancia_pct': _delta_pct(ganancia_neta, prev_ganancia),
            'delta_tickets_pct': _delta_pct(Decimal(tickets), Decimal(prev_tickets)),
        },
        'serie_diaria': daily,
        'top_productos': top_productos,
        'top_categorias': top_categorias,
        'por_sucursal': por_sucursal,
        'por_pago': por_pago,
        'por_estado': por_estado,
        'top_clientes': top_clientes,
        'top_tickets': top_tickets,
    }


def _ganancias_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    _add_logo_to_ws(ws, 'A1')
    bold = Font(bold=True)

    ws.cell(row=8, column=1, value='Reporte de ganancias').font = Font(bold=True, size=14)
    ws.cell(row=9, column=1, value=f'Periodo: {data["periodo_label"]}')
    ws.cell(row=10, column=1, value=f'Desde: {data["desde"][:19]}  Hasta: {data["hasta"][:19]}')

    k = data['kpis']
    summary_rows = [
        ('Tickets', k['tickets']),
        ('Unidades vendidas', k['unidades']),
        ('Ventas (totales)', f'Q {k["ventas_brutas"]}'),
        ('Ingresos brutos por línea', f'Q {k["ingresos"]}'),
        ('Descuentos', f'Q {k["descuentos"]}'),
        ('Costo de mercancía', f'Q {k["costo"]}'),
        ('Ganancia bruta', f'Q {k["ganancia_bruta"]}'),
        ('Ganancia neta', f'Q {k["ganancia_neta"]}'),
        ('Margen', f'{k["margen_pct"]} %'),
        ('Ticket promedio', f'Q {k["ticket_promedio"]}'),
        ('Ganancia promedio por ticket', f'Q {k["ganancia_promedio"]}'),
    ]
    r = 12
    ws.cell(row=r, column=1, value='Indicador').font = bold
    ws.cell(row=r, column=2, value='Valor').font = bold
    r += 1
    for label, value in summary_rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1

    # Comparación
    cmp = data['comparacion']
    r += 2
    ws.cell(row=r, column=1, value='Comparación con periodo anterior').font = bold
    r += 1
    for label, value in [
        ('Ventas previas', f'Q {cmp["ventas_prev"]}'),
        ('Ganancia previa', f'Q {cmp["ganancia_prev"]}'),
        ('Δ Ventas', f'{cmp["delta_ventas_pct"]} %'),
        ('Δ Ganancia', f'{cmp["delta_ganancia_pct"]} %'),
        ('Δ Tickets', f'{cmp["delta_tickets_pct"]} %'),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1

    # Serie diaria
    ws2 = wb.create_sheet('Serie diaria')
    _add_logo_to_ws(ws2, 'A1')
    h = ['Fecha', 'Tickets', 'Unidades', 'Ingresos', 'Costo', 'Descuento', 'Ganancia', 'Margen %']
    for c, t in enumerate(h, start=1):
        ws2.cell(row=8, column=c, value=t).font = bold
    for i, row in enumerate(data['serie_diaria'], start=9):
        ws2.cell(row=i, column=1, value=row['fecha'])
        ws2.cell(row=i, column=2, value=row['tickets'])
        ws2.cell(row=i, column=3, value=row['unidades'])
        ws2.cell(row=i, column=4, value=float(row['ingresos']))
        ws2.cell(row=i, column=5, value=float(row['costo']))
        ws2.cell(row=i, column=6, value=float(row['descuento']))
        ws2.cell(row=i, column=7, value=float(row['ganancia']))
        ws2.cell(row=i, column=8, value=float(row['margen_pct']))

    # Top productos
    ws3 = wb.create_sheet('Top productos')
    _add_logo_to_ws(ws3, 'A1')
    h = ['SKU', 'Producto', 'Categoría', 'Unidades', 'Ingresos', 'Costo', 'Ganancia', 'Margen %', 'Tickets']
    for c, t in enumerate(h, start=1):
        ws3.cell(row=8, column=c, value=t).font = bold
    for i, p in enumerate(data['top_productos'], start=9):
        ws3.cell(row=i, column=1, value=p['sku'])
        ws3.cell(row=i, column=2, value=p['nombre'][:60])
        ws3.cell(row=i, column=3, value=p['categoria'])
        ws3.cell(row=i, column=4, value=p['unidades'])
        ws3.cell(row=i, column=5, value=float(p['ingresos']))
        ws3.cell(row=i, column=6, value=float(p['costo']))
        ws3.cell(row=i, column=7, value=float(p['ganancia']))
        ws3.cell(row=i, column=8, value=float(p['margen_pct']))
        ws3.cell(row=i, column=9, value=p['tickets'])

    # Top categorías
    ws4 = wb.create_sheet('Categorías')
    _add_logo_to_ws(ws4, 'A1')
    h = ['Categoría', 'Productos', 'Unidades', 'Ingresos', 'Costo', 'Ganancia', 'Margen %']
    for c, t in enumerate(h, start=1):
        ws4.cell(row=8, column=c, value=t).font = bold
    for i, c2 in enumerate(data['top_categorias'], start=9):
        ws4.cell(row=i, column=1, value=c2['categoria'])
        ws4.cell(row=i, column=2, value=c2['productos'])
        ws4.cell(row=i, column=3, value=c2['unidades'])
        ws4.cell(row=i, column=4, value=float(c2['ingresos']))
        ws4.cell(row=i, column=5, value=float(c2['costo']))
        ws4.cell(row=i, column=6, value=float(c2['ganancia']))
        ws4.cell(row=i, column=7, value=float(c2['margen_pct']))

    # Sucursales
    ws5 = wb.create_sheet('Sucursales')
    _add_logo_to_ws(ws5, 'A1')
    h = ['Sucursal', 'Tickets', 'Unidades', 'Ingresos', 'Costo', 'Ganancia', 'Margen %']
    for c, t in enumerate(h, start=1):
        ws5.cell(row=8, column=c, value=t).font = bold
    for i, b in enumerate(data['por_sucursal'], start=9):
        ws5.cell(row=i, column=1, value=b['branch_name'])
        ws5.cell(row=i, column=2, value=b['tickets'])
        ws5.cell(row=i, column=3, value=b['unidades'])
        ws5.cell(row=i, column=4, value=float(b['ingresos']))
        ws5.cell(row=i, column=5, value=float(b['costo']))
        ws5.cell(row=i, column=6, value=float(b['ganancia']))
        ws5.cell(row=i, column=7, value=float(b['margen_pct']))

    # Clientes
    ws6 = wb.create_sheet('Clientes')
    _add_logo_to_ws(ws6, 'A1')
    h = ['Cliente', 'NIT', 'Tickets', 'Unidades', 'Ingresos', 'Ganancia']
    for c, t in enumerate(h, start=1):
        ws6.cell(row=8, column=c, value=t).font = bold
    for i, cu in enumerate(data['top_clientes'], start=9):
        ws6.cell(row=i, column=1, value=cu['nombre'])
        ws6.cell(row=i, column=2, value=cu['nit'])
        ws6.cell(row=i, column=3, value=cu['tickets'])
        ws6.cell(row=i, column=4, value=cu['unidades'])
        ws6.cell(row=i, column=5, value=float(cu['ingresos']))
        ws6.cell(row=i, column=6, value=float(cu['ganancia']))

    # Tickets top
    ws7 = wb.create_sheet('Tickets')
    _add_logo_to_ws(ws7, 'A1')
    h = ['Ticket', 'Fecha', 'Cliente', 'Sucursal', 'Líneas', 'Ingresos', 'Descuento', 'Total', 'Ganancia']
    for c, t in enumerate(h, start=1):
        ws7.cell(row=8, column=c, value=t).font = bold
    for i, t in enumerate(data['top_tickets'], start=9):
        ws7.cell(row=i, column=1, value=t['id'])
        ws7.cell(row=i, column=2, value=t['fecha'][:19])
        ws7.cell(row=i, column=3, value=t['cliente'][:50])
        ws7.cell(row=i, column=4, value=t['sucursal'])
        ws7.cell(row=i, column=5, value=t['lineas'])
        ws7.cell(row=i, column=6, value=float(t['ingresos']))
        ws7.cell(row=i, column=7, value=float(t['descuento']))
        ws7.cell(row=i, column=8, value=float(t['total']))
        ws7.cell(row=i, column=9, value=float(t['ganancia']))

    buf = BytesIO()
    wb.save(buf)
    out = buf.getvalue()
    buf.close()
    return out


def _ganancias_pdf(data: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=BOUTIQUE_PDF_PAGE_SIZE, title='Reporte de Ganancias')
    styles = getSampleStyleSheet()
    h_left = ParagraphStyle('gH', parent=styles['Heading2'], alignment=TA_LEFT)
    n_left = ParagraphStyle('gN', parent=styles['Normal'], alignment=TA_LEFT)
    h3_left = ParagraphStyle('gH3', parent=styles['Heading3'], alignment=TA_LEFT)

    logo_buf = pdf_header_image_bytes()
    logo_buf.seek(0)
    with PILImage.open(logo_buf) as pil_im:
        lw, lh = pil_im.size
    logo_buf.seek(0)
    logo_max_w = 96.0
    logo_w = min(logo_max_w, float(lw)) if lw else logo_max_w
    logo_h = logo_w * (float(lh) / float(lw)) if lw else 28.0

    k = data['kpis']
    cmp = data['comparacion']

    story = [
        PdfImage(logo_buf, width=logo_w, height=logo_h, hAlign='LEFT'),
        Spacer(1, 8),
        Paragraph('<b>Reporte de Ganancias</b>', h_left),
        Paragraph(f'Periodo: {data["periodo_label"]}', n_left),
        Paragraph(f'Desde: {data["desde"][:19]} — Hasta: {data["hasta"][:19]}', n_left),
        Spacer(1, 14),
    ]

    # KPIs como tabla
    kpi_rows = [
        ['Tickets', str(k['tickets']), 'Unidades', str(k['unidades'])],
        ['Ventas', f'Q {k["ventas_brutas"]}', 'Ingresos brutos', f'Q {k["ingresos"]}'],
        ['Descuentos', f'Q {k["descuentos"]}', 'Costo', f'Q {k["costo"]}'],
        ['Ganancia bruta', f'Q {k["ganancia_bruta"]}', 'Ganancia neta', f'Q {k["ganancia_neta"]}'],
        ['Margen', f'{k["margen_pct"]} %', 'Ticket promedio', f'Q {k["ticket_promedio"]}'],
        ['Ganancia/ticket', f'Q {k["ganancia_promedio"]}', 'Δ Ganancia', f'{cmp["delta_ganancia_pct"]} %'],
    ]
    kpi_tbl = Table(kpi_rows, colWidths=[110, 120, 110, 120])
    kpi_tbl.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 16))

    # Top productos
    story.append(Paragraph('<b>Top productos por ganancia</b>', h3_left))
    prod_rows = [['SKU', 'Producto', 'Cat.', 'Unid.', 'Ingresos', 'Ganancia', 'Margen %']]
    for p in data['top_productos'][:15]:
        prod_rows.append([
            p['sku'][:14],
            p['nombre'][:32],
            (p['categoria'] or '')[:12],
            str(p['unidades']),
            f'Q {p["ingresos"]}',
            f'Q {p["ganancia"]}',
            f'{p["margen_pct"]}%',
        ])
    prod_tbl = Table(prod_rows, repeatRows=1, colWidths=[60, 150, 60, 40, 64, 64, 50])
    prod_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFF')]),
    ]))
    story.append(prod_tbl)
    story.append(Spacer(1, 14))

    # Sucursales
    story.append(Paragraph('<b>Por sucursal</b>', h3_left))
    suc_rows = [['Sucursal', 'Tickets', 'Unidades', 'Ingresos', 'Ganancia', 'Margen %']]
    for b in data['por_sucursal']:
        suc_rows.append([
            b['branch_name'][:32],
            str(b['tickets']),
            str(b['unidades']),
            f'Q {b["ingresos"]}',
            f'Q {b["ganancia"]}',
            f'{b["margen_pct"]}%',
        ])
    suc_tbl = Table(suc_rows, repeatRows=1, colWidths=[140, 60, 60, 80, 80, 60])
    suc_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFF')]),
    ]))
    story.append(suc_tbl)

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    return pdf


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ganancias_report(request, salida: str | None = None):
    """Reporte de ganancias por semana, quincena o mes.

    Query params:
      - periodo: semana | quincena | mes (default semana)
      - branch:  id de sucursal (opcional)
      - tipo / out / X-Boutique-Report: json | xlsx | pdf
    """
    if salida is not None and str(salida).strip() != '':
        fmt = str(salida).strip().lower()
    else:
        fmt = _report_export_format(request)
    if fmt not in _REPORT_KINDS:
        return _set_no_store(JsonResponse({'detail': f'Formato no válido ({fmt}).'}, status=400))

    periodo = (request.query_params.get('periodo') or 'semana').strip().lower()
    if periodo not in _PERIODOS_GANANCIAS:
        return _set_no_store(JsonResponse(
            {'detail': 'periodo debe ser uno de: semana, quincena, mes.'}, status=400,
        ))
    branch_id = _parse_branch(request)

    data = _ganancias_data(periodo, branch_id)

    if fmt == 'json':
        return _set_no_store(JsonResponse({'generated_at': timezone.now().isoformat(), **data}))

    if fmt == 'xlsx':
        out = _ganancias_xlsx(data)
        resp = HttpResponse(
            out,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = (
            f'attachment; filename="ganancias_{periodo}_{datetime.now():%Y%m%d_%H%M}.xlsx"'
        )
        return _set_no_store(resp)

    if fmt == 'pdf':
        pdf = _ganancias_pdf(data)
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="ganancias_{periodo}_{datetime.now():%Y%m%d_%H%M}.pdf"'
        )
        return _set_no_store(resp)

    return _set_no_store(JsonResponse({'detail': 'Use json, xlsx o pdf.'}, status=400))
